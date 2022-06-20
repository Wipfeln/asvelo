import math
#import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = load_workbook('exlist.xlsx')
#print(wb.sheetnames)
wsratings = wb['rating']
wsscores = wb['tablo']
#p1row= wstable["B"]

RatA = wsratings["B1"].value
RatB = wsratings["B2"].value

Pawin = 1/(1+10**((RatB-RatA)/400))
Pbwin = 1/(1+10**((RatA-RatB)/400))
def rating_calc() :
    global RatA
    global RatB
    RatA = RatA + 32 * (float(ScoreA) - Pawin)
    RatB = RatB + 32 * (float(ScoreB) - Pbwin)
def write_newrat():
    wsratings["C1"] = RatA
    wsratings["C2"] = RatB
    wb.save('exlist.xlsx')
def printall() :
    #print("B=", ScoreB)
    #print("A=", ScoreA)
    print("A=", RatA)
    print("B=", RatB)
ScoreA = wsscores["B5"].value
#ScoreA = input("Enter the Score for A:")
ScoreB = 5

if float(ScoreA) == 1 :
        ScoreB = 0
        rating_calc()
        write_newrat()
        printall()

elif float(ScoreA) == 0 :
        ScoreB = 1
        rating_calc()
        write_newrat()
        printall()
elif float(ScoreA) == 0.5 :
            ScoreB = 0.5
            rating_calc()
            write_newrat()
            printall()

else :
    print("error: Wrong score, please input 1, 0.5, 0 for win, draw, lose respectively")



# for score in p1row:
#     print(score.value)
#print(wstable["B3"].value)
#print(wsratings["A1"].value)

# path = r'C:\Users\lenovo\PycharmProjects\pythonProject\exlist.xlsx'
# df = pd.read_excel(r'C:\Users\lenovo\PycharmProjects\pythonProject\exlist.xlsx')
# print(df)
